import requests
import openpyxl 
from openpyxl.utils import get_column_letter
import re
import sys
#处理周次和课节的数据
def jiejc(a):
    pattern =r'\d+'
    find=re.findall(pattern,a)
    find_numbers=list(map(int,find))#str转为int
    if len(find_numbers)==1:
        return find_numbers
    else:
        b=[]
        for i in range(find_numbers[0],find_numbers[1]+1):
            b.append(i)
        return b
    
try:
    res=requests.post(
        url="http://111.75.254.215:9002/jwglxt/kbcx/xskbcx_cxXsgrkb.html?gnmkdm=N2151&su=2021132697",
        headers={
            "Accept":"*/*",
            "Accept-Encoding":"gzip, deflate",
            "Accept-Language":"zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
            "Content-Length":"28",
            "Content-Type":"application/x-www-form-urlencoded;charset=UTF-8",
            "Cookie":"JSESSIONID=8D93FF3879E5CBF93C4AB396BEC70506",
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.54"
        },
        data={
            "xnm": "2023",
            "xqm": "3",
            "kzlx": "ck",
        }
    ).json()   #zcd  周次，xqjmc 星期，xm 姓名，cdmc 地点，kcmc 课程名称
    r=res["kbList"]
except:
    print(f"没有连接网络或者Cookie过期\n程序终止")
    sys.exit()

# 创建
workbook = openpyxl.Workbook()
worksheet = workbook.active
#设置第一个工作表名字
worksheet.title = "第1周"
for i in range(2,12):
    workbook.create_sheet(f"第{i}周",i)
    
tmokey={"星期一":"B","星期二":"C","星期三":"D","星期四":"E","星期五":"F","星期六":"G","星期日":"H"}
tmo=[["","星期一","星期二","星期三","星期四","星期五","星期六","星期日"],
     ["1"],["2"],["3"],["4"],["5"],["6"],["7"],["8"]]
#在每个工作表写入星期和课节
for worksheet in workbook.worksheets:
    for row in tmo:
        worksheet.append(row)
   
for i in range(len(r)):
    #要写入的数据
    data=f'课名：{r[i]["kcmc"]},老师：{r[i]["xm"]},地点：{r[i]["cdmc"]}'
    for xqjmc in jiejc(r[i]["zcd"]): #周次
        worksheet=workbook.worksheets[xqjmc-1]#选择工作表
        for jc in jiejc(r[i]["jc"]): #课节
            cell = worksheet[f'{tmokey[r[i]["xqjmc"]]}{jc+1}']#选择单元格
            cell.value = data  #进行写入
            #设置宽高
            column_letter = get_column_letter(cell.column)
            worksheet.column_dimensions[column_letter].width = 65
name=res["xsxx"]
workbook.save(f'{name["BJMC"]}{name["XM"]}.xlsx')
print("导出成功")