import openpyxl 
import re
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.utils import get_column_letter
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import requests
import os
import sys
 
 #要连接网络，最简单的检查网络的方法就是ping了
ret = os.system("ping baidu.com -n 1")
if ret != 0:
    print("没有连接网络，程序即将退出")
    time.sleep(3)
    sys.exit()

#处理单双周次和课节的数据
def jiejc(a):
    pattern =r'\d+'
    find=re.findall(pattern,a)
    find_numbers=list(map(int,find))#str转为int
    d=a.split(",")
    if "周" in a:
        c=[]
        for week in d:
            if "-" in week:
                start, end = re.findall(r"\d+", week)
                start = int(start)
                end = int(end)
                # 处理单周或双周
                if "单" in week:
                    for k in range(start,end+1):
                        if k%2!=0:
                            c.append(k)
                elif "双" in week:
                    for j in range(start,end+1):
                        if j%2==0:
                            c.append(j) 
                else:
                    for i in range(start, end + 1):
                        c.append(i)
            else:
                week_num = int(re.findall(r"\d+", week)[0])
                c.append(week_num)
        return c
    else:
        b=[]
        for i in range(find_numbers[0],find_numbers[1]+1,2):
            b.append(f'{i}-{i+1}节')
        return b
    
#-------------通过江西应用职业技术学院的教务系统获取课表数据--------
#通过学号和年份就可以获取个人课表（任何专业都可）
user=input('请输入学号：')
xnm=input('请输入年份(如:2023):')
xq=input('请输入学期(1-2):')
if xq=='1':
    xqm='3'
if xq=='2':
    xqm='12'

driver = webdriver.Edge()

# 打开登录页面
driver.get('https://jw.jxyy.edu.cn:9002/jwglxt/xtgl/login_slogin.html')
# 执行登录操作，输入用户名和密码
username_input = driver.find_element(By.ID, "yhm")
password_input = driver.find_element(By.ID, "mm")
username_input.send_keys(user)
password_input.send_keys('xs123456')

# 点击登录按钮
login_button = driver.find_element(By.ID, "dl")
login_button.click()

# 等待登录完成和加载Cookie
driver.implicitly_wait(10)

time.sleep(3)
# 获取所有的Cookie信息
cookies = driver.get_cookies()

# 关闭浏览器
driver.quit()
aa=cookies[0]
cok=f'{aa["name"]}={aa["value"]}'
res=requests.post(
    url=f"https://jw.jxyy.edu.cn:9002/jwglxt/kbcx/xskbcx_cxXsgrkb.html?gnmkdm=N2151&su={user}",
    headers={
        "Accept":"*/*",
        "Accept-Encoding":"gzip, deflate",
        "Accept-Language":"zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "Content-Length":"28",
        "Content-Type":"application/x-www-form-urlencoded;charset=UTF-8",
        "Cookie":cok,
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.54"
    },
    data={
        "xnm": xnm,
        "xqm": xq,
        "kzlx": "ck",
    }
).json()
r=res["kbList"]

# 创建
workbook = openpyxl.Workbook()
worksheet = workbook.active
#创建表
tom=["","",""]
ton=["星期一","星期二","星期三","星期四","星期五","星期六","星期日"]
kj=["1-2节","3-4节","5-6节","7-8节","9-10节"]
for r1 in range(1,22):
    tom.append(f'第{r1}周')
worksheet.append(tom)
               
for j in range(7):
        # 合并单元格  A列
        worksheet.merge_cells(f'A{2+j*5}:A{6+j*5}')
        # 设置合并后的值
        merged_value = ton[j]
        worksheet[f'A{2+j*5}'] = merged_value
        #B列
        worksheet.merge_cells(f'B{2+j*5}:B{3+j*5}')
        worksheet[f'B{2+j*5}'] = "上午"
        worksheet.merge_cells(f'B{4+j*5}:B{5+j*5}')
        worksheet[f'B{4+j*5}'] = "下午"
        worksheet[f'B{6+j*5}'] = "晚上"
        #C列
        for ki in range(len(kj)):
            worksheet[f'C{2+j*5+ki}'] = kj[ki]
            
for i in range(1,4):#纵
    worksheet.column_dimensions[get_column_letter(i)].width = 4

for row in worksheet.iter_rows():
   for cell in row:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True,horizontal='center', vertical='center')
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                           right=openpyxl.styles.Side(style='thin'),
                                           top=openpyxl.styles.Side(style='thin'),
                                           bottom=openpyxl.styles.Side(style='thin'))
        cell.font = openpyxl.styles.Font(name='宋体', size=10)

kem=[]
for i in range(len(r)):
    
    for k in jiejc(r[i]["zcd"]):
        a1=worksheet[1]
        for row in a1:
            if row.value == f'第{k}周':
                zcd=row.column_letter
        for j in jiejc(r[i]["jc"]):
            b=0
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == j:
                        b=b+1
                        if (b==int(r[i]["xqj"])):
                            xqj=cell.row
                            dan=f'{zcd}{xqj}'
                            data=f'课名：{r[i]["kcmc"]},老师：{r[i]["xm"]},地点：{r[i]["cdmc"]}'
                            kem.append(r[i]["kcmc"])
                            print(data)#高逼格
                            cell = worksheet[dan]#选择单元格
                            cell.value = data  #进行写入
                            
#单元格规制写入
color=["FFFFCC","CCFFFF","FFCCCC","99CCCC","FFCC99","FFFF99","CCCCFF","FF6666","666666","CCCCFF","003399","FF9900","CC6699","CCFF66","99CCCC"]
myList = list(set(kem))
for ad in range(len(myList)):
    red_fill = PatternFill(bgColor=color[ad])
    dxf = DifferentialStyle(fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text=myList[ad], dxf=dxf)
    rule.formula = [f'NOT(ISERROR(SEARCH("{myList[ad]}",A1)))']
    worksheet.conditional_formatting.add('A1:X40', rule)
#设置宽高
for i in range(2, worksheet.max_row+1):#横
    worksheet.row_dimensions[i].height = 50
for i in range(4, worksheet.max_column+1):#纵
    worksheet.column_dimensions[get_column_letter(i)].width = 17
workbook.save(f'{xnm}年第{xq}学期{res["xsxx"]["BJMC"]}{res["xsxx"]["XM"]}课表.xlsx')
# 关闭Excel文件
workbook.close()
print("导出成功")